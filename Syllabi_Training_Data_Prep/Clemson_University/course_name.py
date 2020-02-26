"""
    Get File Name from parsed text for each course
"""
import re
import pandas as pd

from openpyxl import load_workbook

READ_EXCEL_FILE = '/Users/cheng/Desktop/Testing/Fall_2018_Results.xlsm'
WRITE_EXCEL_FILE = '/Users/cheng/Desktop/Testing/T.xlsx'

# Get File Name based on comparison between exsiting titles and strings
def get_course_name(inputfile, outputfile, titlefile=None):
    """
    Keyword Arguments:
        inputfile {str} -- path of file to parse
        outputfile {str} -- path of file to save
        titlefile {str} -- path of file with all titles (default: {None})
    """

    if titlefile is None:
        titlefile = '/Users/cheng/Desktop/Testing/Course_Title.xlsm'

    title = pd.DataFrame(pd.read_excel(titlefile, keep_default_na=False))
    data = pd.DataFrame(pd.read_excel(inputfile, keep_default_na=False))
    excelfile = load_workbook(outputfile)
    excelfile1 = excelfile.active
    titles = []

    # get all titles:
    for i in range(title.iloc[:, 0].size):
        buff = title.loc[i][0]
        buff = str(buff)
        titles.append(buff)

    for i in range(len(data['file_beginning'])):
        text = ''
        length = 0
        info = data.loc[i][1]
        info = str(info)

        for title in titles:
            if title.lower() in info.lower():
                # useually the longest match string is the title for that course
                if len(title) > length:
                    text = title
                    length = len(title)

        excelfile1.cell(i+2, 1, text)

    excelfile.save(outputfile)
    print('reday')

# get ID from file full-name
def get_id(inputfile, outputfile):
    """
    Keyword Arguments:
        inputfile {str} -- path of file to parse
        outputfile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(inputfile, keep_default_na=False))
    excelfile = load_workbook(outputfile)
    excelfile1 = excelfile.active

    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][0]

        if buff:
            buff = str(buff)
            numbers = re.findall(r'\d+', buff)
            text = ''

            for number in numbers:
                num = int(number)
                if (num >= 100 and num <= 9999 and num != 2015 and
                        num != 2016 and num != 2017 and num != 2018 and num != 2019):
                    text = number
                    break

            excelfile1.cell(i+2, 1, text)

    excelfile.save(outputfile)
    print('ready')

# Deal with whole beginning
def get_course_title_beginning(inputfile, outputfile):
    """
    Keyword Arguments:
        inputfile {str} -- path of file to parse
        outputfile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(inputfile, keep_default_na=False))
    excelfile = load_workbook(outputfile)
    excelfile1 = excelfile.active

    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][2]
        number = data.loc[i][1]
        num = str(number)
        text = ""

        # Find the course title based on course ID, focus on three lines only
        if buff:
            buff = str(buff)
            words = re.split('[\n]', buff)
            final_words = []

            for word in words:
                if word.strip() != '':
                    final_words.append(word)

            if num == '':
                text = buff
            elif num not in buff:
                text = buff
            else:
                for j in range(len(final_words)):
                    if num in final_words[j]:
                        if len(final_words[j]) < 200:
                            text = final_words[j]

                        if j == 0 and len(final_words) > 2:
                            text = text + '\n' + final_words[j + 1]
                            text = text + '\n' + final_words[j + 2]

                        if j > 0:
                            text = text + '\n' + final_words[j - 1]

                        if j < len(final_words) - 1:
                            text = text + '\n' + final_words[j + 1]

                        if j == len(final_words) - 1 and len(final_words) > 2:
                            text = text + '\n' + final_words[j - 1]
                            text = text + '\n' + final_words[j - 2]
                        break

        text = text.replace('\t', ' ')
        excelfile1.cell(i+2, 3, text)

    excelfile.save(outputfile)
    print('ready')

# remove unrelated parts
def remove_parts_from_beginning(inputfile, outputfile):
    """
    Keyword Arguments:
        inputfile {str} -- path of file to parse
        outputfile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(inputfile, keep_default_na=False))
    excelfile = load_workbook(outputfile)
    excelfile1 = excelfile.active

    for i in range(len(data['file_beginning'])):
        text = data.loc[i][3]

        if text:
            text = str(text)

        number = data.loc[i][1]
        num = str(number)

        if num in text:    # remove course ID
            text = re.sub(r'\&?\s?[A-Z][A-Z,a-z,\/]{1,10}\s+[A-Z]?[\d,\/]{3,9}\:?', "", text)

        if ':' in text:    # remove time
            text = re.sub(r'\d{1,2}\:\d{2}', "", text)

        if '.' in text:    # remove section ID
            text = re.sub(r'\.\d{2,3}', "", text)

        if '-' in text:    # remove section ID
            text = re.sub(r'\-\d{2,3}', "", text)

        if '/' in text:    # remove section ID
            text = re.sub(r'\/\d{2,4}', "", text)

        if '‚Äô' in text:    # replace odd signs
            text = text.replace('‚Äô', "‘")

        if '¬†' in text:    # remove odd signs
            text = text.replace('¬†', " ")

        if 'Syllabus for' in text:    # remove unrelated info
            text = re.sub(r'^Syllabus for [0-9]{1,3}', "", text)

        if 'and' in text:    # remove unrelated info
            text = re.sub(r'\band [0-9]{1,4}', "", text)

        if 'Fall' in text:    # remove semester
            text = re.sub(r'\bFall [0-9]{4}', "", text)

        if 'Sections' in text:    # remove section ID
            text = re.sub(r'\bSections [0-9]{1,3}', "", text)

        if 'Section' in text:    # remove section ID
            text = re.sub(r'\bSection [0-9]{1,3}', "", text)

        if 'section' in text:    # remove section ID
            text = re.sub(r'\bsection [0-9]{1,3}', "", text)

        if 'Sec.' in text:    # remove section ID
            text = re.sub(r'\bSec. [0-9]{1,3}', "", text)

        if 'sec.' in text:    # remove section ID
            text = re.sub(r'\bsec. [0-9]{1,3}', "", text)

        if 'Dr.' in text:    # remove instructor's name
            text = re.sub(r'\bDr. [A-Z][a-z]{2,16}\s{1}[A-Z][a-z]{2,16}', "", text)

        # Remove unrelated info
        words = ['Credit Hours', 'Credit Hour', 'Variable credits', 'credit', 'and RCID',
                 'GENERAL INFORMATION', 'General Information', 'Äê', 'Äù', 'Äî', 'Ä¢',
                 'Äú', 'àí', 'Äì', 'Äò', '¬≠', '†', '≠ ê', 'Æ', '¬', 'TR ', 'pm ', '[',
                 ']', '_', 'Mon.', 'Wed.', 'Fri.', 'CI Team', 'TTh', 'Fall One', '(',
                 'Fall Two', 'Mini A', '-Fall', 'ONLINE', 'Session', 'MINIMESTER A', ';',
                 'Minimester C', 'Mini C', 'Minimester A', 'FALL II', 'FALL I', 'FALL',
                 'SYLLABUS', 'Full Term', 'Fall Minimester A', 'Course Syllabus', 'Spring',
                 '¬†', 'COURSE OUTLINE', 'Syllabus', 'Fall II', 'Fall I', 'Fall', ')',
                 'Course Information and Policies (General Syllabus)', '2015', '2016',
                 '2017', '2018', '2019', 'CLEMSON UNIVERSITY', 'Clemson University']
        for word in words:
            if word in text:
                text = text.replace(word, "")

        if '.' in text:
            text = text.replace('.', " ")

        excelfile1.cell(i+2, 5, text)

    excelfile.save(outputfile)
    print('ready')

# remove lines
def remove_lines_from_beginning(inputfile, outputfile):
    """
    Keyword Arguments:
        inputfile {str} -- path of file to parse
        outputfile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(inputfile, keep_default_na=False))
    excelfile = load_workbook(outputfile)
    excelfile1 = excelfile.active

    # Mark the unwanted lines
    del_words = ['MSON UNIVERSITY', 'auditorium', 'about me', 'a.m.', 'p.m.', 'Tues.', 'Thur.',
                 'Tue ', 'Thu ', 'LEGE ', 'Mr.', 'Ms.', 'building', 'LEGE ', 'weekly', 'LEGE ',
                 'TBD', 'son University', 'manual', 'time', 'session', 'summer', 'credits',
                 'credit hr', 'credit units', 'schedule', 'dates', 'http', 'cell', 'lecture',
                 'meeting times', 'duration', 'advisor', 'coordinator', 'instructor', 'center',
                 'professor', 'course outline', 'requirements', 'course calendar', 'lecturer',
                 'prof ', 'm/w', 'teaching assistant', 't, th', 'mon/wed/fri', 't/th', 'm.w.f.',
                 'tu & th', 'tu ', 'm, w, f', 'wed.', 'mw:', 'semester', 'clemson university ',
                 'ph.d', 'prof.', 'room ', 'tth', 'hours', 'office', 'address', 'phone', 'hall',
                 'department', 'school', 'college', 'page', '@', 'august', 'mwf', 'monday',
                 'friday', 'tuesday', 'thursday', 'wednesday', 'location', 'tth ', 'contact ',
                 'Credit hour', 'credit', 'am- pm', 'Labs are in']

    # Mark the add back lines
    add_words = ['multiple sections', 'Cancer Cell Comparisons', 'General Chemistry 1st Semester',
                 'Thesis Hours', 'High School', 'Public School', 'Elementary School',
                 'Primary School', 'School Counseling', 'School  Administrators',
                 'School Administration', 'PROFESSIONAL', 'Introduction to the Course',
                 'Middle School Curriculum']

    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][4]
        text = ""
        if buff:
            buff = str(buff)
            words = re.split('[\n]', buff)
            for word in words:
                if len(word.strip()) > 4:
                    for delw in del_words:
                        if delw in word.lower():
                            break
                    else:
                        text = text + '\n' + word.strip()

                    for addw in add_words:
                        if addw in word:
                            text = text + '\n' + word.strip()

                else:
                    continue

        excelfile1.cell(i+2, 7, text)

    excelfile.save(outputfile)
    print('ready')

if __name__ == "__main__":
    remove_lines_from_beginning(READ_EXCEL_FILE, WRITE_EXCEL_FILE)
#     remove_parts_from_beginning(READ_EXCEL_FILE, WRITE_EXCEL_FILE)
#     get_course_title_beginning(READ_EXCEL_FILE, WRITE_EXCEL_FILE)
#     get_id(READ_EXCEL_FILE, WRITE_EXCEL_FILE)
#     get_course_name(READ_EXCEL_FILE, WRITE_EXCEL_FILE)

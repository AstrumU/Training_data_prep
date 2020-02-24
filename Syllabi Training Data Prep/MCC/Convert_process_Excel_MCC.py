"""
   Get Useful Content
"""
import csv
import re
import string
import xlwt
import datetime
import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook

# File Path {str}
read_excel_file = '/Users/cheng/Desktop/astrumu_parsers-master/tests/test_files/Courses/output/batch_process_result.csv'
write_excel_file = '/Users/cheng/Desktop/Testing/Schema.xlsx'

# All Possible Markers
Certified_Markers = {'DATE SUBMITTED','CATALOG NO.','DATE DICC APPROVED','DATE LAST REVIEWED','COURSE INFORMATION FORM','DISCIPLINE', 
                     'COURSE TITLE','CR.HR','LECT HR.','LAB HR.','CLIN/INTERN HR.','CLOCK HR.','CATALOG DESCRIPTION','PREREQUISITES',
                     'EXPECTED STUDENT OUTCOMES IN THE COURSE','GENERAL EDUCATION OUTCOMES','PROGRAM-LEVEL OUTCOMES',
                     'CAREER AND TECHNICAL EDUCATION PROGRAM OUTCOMES','CLASS-LEVEL ASSESSMENT MEASURES','COURSE OUTLINE FORM',
                     'PROGRAM-LEVEL OUTCOMES ADDRESSED','COURSE INFORMATION','COURSE OUTLINE','DATE DICC','DICC APPROVAL','NO.',
                     'PROGRAM OUTCOMES','ASSESSMENT MEASURES'}

# Read in input csv document and activate output xlsx document
df = pd.read_csv(read_excel_file)   
wb = load_workbook(write_excel_file)
wb1 = wb.active

# Regular Expression
p = r'((^|\t|\n){0,}[A-Z.]+[ \-\/\n\t]{0,1}){1,}(([A-Z()]|\.){2,}(\t|\s{4}|\n\$){0,})'
pattern = re.compile(p)

# Iterate Over the CSV Document
for i in range(len(df['file_fullname'])):
    # Copy the file name into output document
    wb1.cell(i+2,1,df['file_fullname'][i])

    # Read in and process file text
    buff = df['file_text'][i]
    if isinstance(buff, str):
        buff = buff.replace("|","")      # Some parsed results have unwanted symbols
        words = re.split('[\n\t]', buff)      # Break down file text
        markers = []

        for word in words:
            for uppercase in re.finditer(pattern, word):

                marker = uppercase[0]
                marker = marker.strip()

                if marker in Certified_Markers:
                    markers.append(marker)

                """
                    Special Cases 1: Two markers are put together
                """

                if marker == 'COURSE INFORMATION FORM DISCIPLINE':
                    markers.append('COURSE INFORMATION FORM')                    
                    markers.append('DISCIPLINE')
                if marker == 'COURSE OUTLINE FORM DISCIPLINE':
                    markers.append('COURSE OUTLINE FORM')                    
                    markers.append('DISCIPLINE') 
                if marker == 'CLOCK HR. CATALOG DESCRIPTION':
                    markers.append('CLOCK HR.')
                    markers.append('CATALOG DESCRIPTION')

                """
                    Special Cases 2: Remove redundant '(ESO)'
                """
                
                if marker == 'EXPECTED STUDENT OUTCOMES IN THE COURSE (ESO)':
                    markers.append('EXPECTED STUDENT OUTCOMES IN THE COURSE')
                if marker == 'GENERAL EDUCATION OUTCOMES (ESO)':
                    markers.append('GENERAL EDUCATION OUTCOMES')
                if marker == 'IN THE COURSE (ESO)':
                    markers.append('IN THE COURSE')
                if marker == 'OUTCOMES (ESO)':
                    markers.append('OUTCOMES')
                    
        markers.append('QWERTYUIOP')      # Add an extra unique marker to find the end of the text
        
        for j in range(len(markers) -1):
            buff = buff + "QWERTYUIOP"      # Append string corresponding to the unique marker

            """
                This is the last part. Some duplicate markers in this part need not to be treated as markers
            """
            if markers[j] == 'COURSE OUTLINE FORM' or markers[j] == 'COURSE OUTLINE':
                pattern0 = re.compile(markers[j] + '(.*?)' + markers[-1],re.S)
                result = pattern0.findall(buff)
                if result:
                    wb1.cell(i+2,20,result[0].strip())
                    break
            else:
                pattern0 = re.compile(markers[j] + '(.*?)' + markers[j+1],re.S)
                result = pattern0.findall(buff)

            """
                Put the content under the corresponding marker
            """

            if result:
                if markers[j] == 'DATE SUBMITTED':
                    wb1.cell(i+2,2,result[0].strip())
                elif markers[j] == 'CATALOG NO.' or markers[j] == 'NO.':
                    wb1.cell(i+2,3,result[0].strip())
                elif markers[j] == 'DATE DICC APPROVED' or markers[j] == 'DATE DICC' or markers[j] == 'DICC APPROVAL':
                    wb1.cell(i+2,4,result[0].strip())
                elif markers[j] == 'DATE LAST REVIEWED':
                    wb1.cell(i+2,5,result[0].strip())
                elif markers[j] == 'DISCIPLINE':
                    wb1.cell(i+2,6,result[0].strip())
                elif markers[j] == 'COURSE TITLE':
                    wb1.cell(i+2,7,result[0].strip())                
                elif markers[j] == 'CR.HR':
                    wb1.cell(i+2,8,result[0].strip())
                elif markers[j] == 'LECT HR.':
                    wb1.cell(i+2,9,result[0].strip())           
                elif markers[j] == 'LAB HR.':
                    wb1.cell(i+2,10,result[0].strip())                
                elif markers[j] == 'CLIN/INTERN HR.':
                    wb1.cell(i+2,11,result[0].strip())
                elif markers[j] == 'CLOCK HR.':
                    wb1.cell(i+2,12,result[0].strip())
                elif markers[j] == 'CATALOG DESCRIPTION':
                    wb1.cell(i+2,13,result[0].strip())
                elif markers[j] == 'PREREQUISITES':
                    wb1.cell(i+2,14,result[0].strip())


                # Get rid of '(ESO)' if there is any
                elif markers[j] == 'EXPECTED STUDENT OUTCOMES IN THE COURSE' or markers[j] == 'IN THE COURSE':
                    temporary = result[0].strip()
                    if temporary[:5] == '(ESO)':
                        wb1.cell(i+2,15,temporary[6:])
                    else:
                        wb1.cell(i+2,15,temporary)
                elif markers[j] == 'GENERAL EDUCATION OUTCOMES' or markers[j] == 'OUTCOMES':
                    temporary = result[0].strip()
                    if temporary[:5] == '(ESO)':
                        wb1.cell(i+2,16,temporary[6:])
                    else:
                        wb1.cell(i+2,16,temporary)

                elif markers[j] == 'CAREER AND TECHNICAL EDUCATION PROGRAM OUTCOMES' or markers[j] == 'PROGRAM OUTCOMES':
                    wb1.cell(i+2,17,result[0].strip())
                elif markers[j] == 'CLASS-LEVEL ASSESSMENT MEASURES' or markers[j] == 'ASSESSMENT MEASURES':
                    wb1.cell(i+2,18,result[0].strip())
                elif markers[j] == 'PROGRAM-LEVEL OUTCOMES ADDRESSED':
                    wb1.cell(i+2,19,result[0].strip())
                else:
                    continue

            wb1.cell(i+2,21,datetime.datetime.now())
            
            """
                Extract the remaing text
            """  
            pattern1 = re.compile(markers[j]+'(.*?)'+markers[-1],re.S)
            temp = pattern1.findall(buff)
            buff = temp[0]
    else:
        continue

"""
    Save the document
""" 
wb.save(write_excel_file)
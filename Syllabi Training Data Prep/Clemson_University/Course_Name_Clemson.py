import csv
import re
import string
import xlwt
import datetime
import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook

read_excel_file = '/Users/cheng/Desktop/Testing/Fall_2018_Results.xlsm'
write_excel_file = '/Users/cheng/Desktop/Testing/T.xlsx'

InputFile = read_excel_file
OutputFile = write_excel_file

# Get File Name based on comparison between exsiting titles and strings
def Get_Course_Name(InputFile, OutputFile, TitleFile = None):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
        TitleFile {str} -- path of file with all titles (default: {None})
    """

    if TitleFile is None:
        TitleFile = '/Users/cheng/Desktop/Testing/Course_Title.xlsm'

    title = pd.DataFrame(pd.read_excel(TitleFile, keep_default_na=False))
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active
    Titles = []

    # get all titles:
    for i in range(title.iloc[:,0].size):
        buff = title.loc[i][0]
        buff = str(buff)
        Titles.append(buff)
        
    for i in range(len(data['file_beginning'])):
        text = ''
        length = 0
        info = data.loc[i][1]
        info = str(info)    
        
        for title in Titles:
            if title.lower() in info.lower():
                # useually the longest match string is the title for that course
                if len(title) > length:
                    text = title
                    length = len(title)
        
        wb1.cell(i+2,1,text)
      
    wb.save(write_excel_file)
    print('reday')


# get ID from file full-name
def Get_ID(InputFile, OutputFile):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active

    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][0]
        
        if buff:
            buff = str(buff)
            numbers = re.findall(r'\d+', buff)
            text = ''
            
            for number in numbers:
                num = int(number)
                if num >= 100 and num <= 9999 and num != 2015 and num != 2016 and num != 2017 and num != 2018 and num != 2019:
                    text = number
                    break

            wb1.cell(i+2,1,text)
            
    wb.save(write_excel_file)
    print('ready')

# Deal with whole beginning
def Get_Course_Title_beginning(InputFile, OutputFile):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active
    
    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][2]
        number = data.loc[i][1]
        num = str(number)
        text = ""
        
        # Find the course title based on course ID, focus on three lines only
        if buff:
            buff = str(buff)
            words = re.split('[\n]', buff)
            final_words = []

            for word in words:
                if word.strip() != '':
                    final_words.append(word)       

            if num == '':
                text = buff
            elif num not in buff:
                text = buff
            else:
                for j in range(len(final_words)):
                    if num in final_words[j]:
                        if len(final_words[j]) < 200:
                            text = final_words[j]
                        
                        if j == 0 and len(final_words) > 2:
                            text = text + '\n' + final_words[j + 1]
                            text = text + '\n' + final_words[j + 2]
                        
                        if j > 0:
                            text = text + '\n' + final_words[j - 1]

                        if j < len(final_words) - 1:
                            text = text + '\n' + final_words[j + 1]
                            
                        if j == len(final_words) - 1 and len(final_words) > 2:
                            text = text + '\n' + final_words[j - 1]
                            text = text + '\n' + final_words[j - 2]                        
                        break 

        text = text.replace('\t', ' ')
        wb1.cell(i+2,3,text)

    wb.save(write_excel_file)
    print('ready')
    
# remove unrelated parts
def Remove_parts_from_beginning(InputFile, OutputFile):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active
    
    for i in range(len(data['file_beginning'])):
        text = data.loc[i][3]

        if text:
            text = str(text)

        number = data.loc[i][1]
        num = str(number)

        if num in text:    # remove course ID
            text = re.sub(r'\&?\s?[A-Z][A-Z,a-z,\/]{1,10}\s+[A-Z]?[\d,\/]{3,9}\:?', "", text)   # ECE 8930 007:, English 214, EDF9790, EdF 4800

        if ':' in text:    # remove time
            text = re.sub(r'\d{1,2}\:\d{2}', "", text)
            
        if '.' in text:    # remove section ID
            text = re.sub(r'\.\d{2,3}', "", text)       

        if '-' in text:    # remove section ID
            text = re.sub(r'\-\d{2,3}', "", text)
        
        if '/' in text:    # remove section ID
            text = re.sub(r'\/\d{2,4}', "", text)
            
        if '‚Äô' in text:    # replace odd signs
            text = text.replace('‚Äô', "‘")
            
        if '¬†' in text:    # remove odd signs
            text = text.replace('¬†', " ")
                    
        if 'Syllabus for' in text:    # remove unrelated info
            text = re.sub(r'^Syllabus for [0-9]{1,3}', "", text)
        
        if 'and' in text:    # remove unrelated info    
            text = re.sub(r'\band [0-9]{1,4}', "", text)
        
        if 'Fall' in text:    # remove semester         
            text = re.sub(r'\bFall [0-9]{4}', "", text)
        
        if 'Sections' in text:    # remove section ID
            text = re.sub(r'\bSections [0-9]{1,3}', "", text)            
        
        if 'Section' in text:    # remove section ID
            text = re.sub(r'\bSection [0-9]{1,3}', "", text)
            
        if 'section' in text:    # remove section ID
            text = re.sub(r'\bsection [0-9]{1,3}', "", text)
            
        if 'Sec.' in text:    # remove section ID
            text = re.sub(r'\bSec. [0-9]{1,3}', "", text)
            
        if 'sec.' in text:    # remove section ID
            text = re.sub(r'\bsec. [0-9]{1,3}', "", text)        
        
        if 'Dr.' in text:    # remove instructor's name         
            text = re.sub(r'\bDr. [A-Z][a-z]{2,16}\s{1}[A-Z][a-z]{2,16}', "", text)                                  
    
        """
            Remove unrelated info
        """
        words = ['Credit Hours', 'Credit Hour', 'Variable credits', 'credits', 'and RCID', 'GENERAL INFORMATION', 'General Information', 
                 'Äê', 'Äù', 'Äî','Ä¢', 'Äú', 'àí', 'Äì', 'Äò','¬≠', '†', '≠ ê', 'Æ', '¬', 'TR ', 'pm ', '[', ']', '_', 'Mon.', 'Wed.', 'Fri.', 
                 'CI Team', 'TTh', 'Fall One', 'Fall Two', 'Mini A', '-Fall', 'ONLINE', 'Session', 'MINIMESTER A', 'Minimester C', 'Mini C', 'Minimester A', 
                 'FALL II', 'FALL I', 'FALL', 'SYLLABUS', 'Full Term', 'Fall Minimester A', 'Course Syllabus', 'Spring', ';', '¬†',
                 'COURSE OUTLINE', 'Syllabus', 'Fall II','Fall I', 'Fall', 'Course Information and Policies (General Syllabus)', 
                 '2015', '2016', '2017', '2018', '2019', 'CLEMSON UNIVERSITY', 'Clemson University', '(', ')']
        for word in words:                                                                       
            if word in text:                           
                text = text.replace(word, "")
                
        if '.' in text:
            text = text.replace('.', " ")        
       
        wb1.cell(i+2,5,text)          

    wb.save(write_excel_file)
    print('ready')
    
# remove lines
def Remove_lines_from_beginning(InputFile, OutputFile):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active   
    
    # Mark the unwanted lines
    del_words = ['MSON UNIVERSITY', 'auditorium', 'about me', 'a.m.', 'p.m.', 'Tues.', 'Thur.', 'Tue ', 'Thu ', 
                 'LEGE ', 'Mr.', 'Ms.', 'building', 'LEGE ', 'weekly', 'LEGE ', 'TBD', 'son University', 'manual', 
                 'time', 'session', 'summer', 'credits', 'credit hr', 'credit units', 'schedule', 'dates', 'http', 
                 'cell', 'lecture', 'meeting times', 'duration', 'advisor', 'coordinator', 'instructor', 'center', 
                 'professor', 'course outline', 'requirements', 'course calendar', 'lecturer', 'prof ', 'm/w', 
                 'teaching assistant', 't, th', 'mon/wed/fri', 't/th', 'm.w.f.', 'tu & th', 'tu ', 'm, w, f', 'wed.', 
                 'mw:', 'semester', 'clemson university ', 'ph.d', 'prof.', 'room ', 'tth', 'hours', 'office', 'address', 
                 'phone', 'hall', 'department', 'school', 'college', 'page', '@', 'august', 'mwf', 'monday', 'friday', 
                 'tuesday', 'thursday', 'wednesday', 'location', 'tth ', 'contact ', 'Credit hour', 'credit', 'am- pm', 'Labs are in']
    
    # Mark the add back lines
    add_words = ['multiple sections', 'Cancer Cell Comparisons', 'General Chemistry 1st Semester', 'Thesis Hours', 
                 'High School', 'Public School', 'Elementary School', 'Primary School', 'School Counseling', 
                 'School  Administrators', 'School Administration', 'PROFESSIONAL', 'Introduction to the Course', 
                 'Middle School Curriculum']
    
    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][4]
        text = ""                                                                    
        if buff:
            buff = str(buff)
            words = re.split('[\n]', buff)              
            for word in words:
                if len(word.strip()) > 4:
                    for delw in del_words:
                        if delw in word.lower():
                            break
                    else:
                        text = text + '\n' + word.strip()
                            
                    for addw in add_words:
                        if addw in word:
                            text = text + '\n' + word.strip()

                else:             
                    continue                                       
                   
        wb1.cell(i+2,7,text)

    wb.save(write_excel_file)
    print('ready')
    
if __name__ == "__main__":
    Remove_rows_from_beginning(InputFile, OutputFile)
#     Remove_parts_from_beginning(InputFile, OutputFile)
#     Get_Course_Title_beginning(InputFile, OutputFile)
#     Get_ID(InputFile, OutputFile)
#     Get_Course_Name(InputFile, OutputFile)